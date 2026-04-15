# -*- coding: utf-8 -*-
import io
import re
import datetime as dt

import pandas as pd
import streamlit as st

col1, col2 = st.columns([8, 2])


with col2:
    with open("arquivos_padrao/comparacao_ciclo_padrao.xlsx", "rb") as file:
        st.download_button(
            label="📥 Baixar arquivo padrão",
            data=file,
            file_name="comparacao_ciclo_padrao.xlsx.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# === SIDEBAR (IGUAL AO HOME) ===
st.markdown(
    """
    <style>
        [data-testid="stSidebarNav"] {
            display: none;
        }
    </style>
    """,
    unsafe_allow_html=True
)

st.sidebar.image("images/agco.jpg", width=240)
st.sidebar.divider()

st.sidebar.markdown("### 📊 Aplicações")
st.sidebar.page_link("Home.py", label="🏠 Home")
st.sidebar.page_link("pages/1_Nivelamento.py", label="📈 Nivelamento sem filas")
st.sidebar.page_link("pages/2_NIvelar_com_Filas.py", label="🛠 Nivelamento com Filas")
st.sidebar.page_link("pages/3_Comparacao_Ciclo.py", label="🔄 Comparativo ciclo Demand Review")
# =================================

# =====================================================
# CONFIGURAÇÃO DA PÁGINA
# =====================================================
st.set_page_config(
    page_title="Comparativo Request Vs Plan",
    layout="wide"
)

st.title("Comparativo Request Vs Plan")
st.caption("Comparativo entre cenários com filtros, resumos e exportação")

PT_BR_MESES = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]

MES_RE = re.compile(
    r'^(jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)/\d{2}$',
    re.IGNORECASE
)

# =====================================================
# FUNÇÕES UTILITÁRIAS
# =====================================================
def _normalize_header(col):
    if isinstance(col, (pd.Timestamp, dt.date)):
        return f"{PT_BR_MESES[col.month - 1]}/{col.year % 100:02d}"

    s = str(col).replace("\u00a0", " ").strip().lower()
    s = re.sub(r"[-_ ]+", "/", s)

    m = re.match(r"^(jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)/(\d{2,4})$", s)
    if m:
        return f"{m.group(1)}/{m.group(2)[-2:]}"

    m = re.match(r"^(jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)(\d{2})$", s)
    if m:
        return f"{m.group(1)}/{m.group(2)}"

    return s


def detectar_colunas_mes(df):
    cols_mes = []
    debug_map = {}

    for c in df.columns:
        alias = _normalize_header(c)
        debug_map[str(c)] = alias
        if MES_RE.match(alias or ""):
            cols_mes.append(c)

    def ordem(c):
        mm, yy = debug_map[str(c)].split("/")
        return int(yy), PT_BR_MESES.index(mm)

    return sorted(cols_mes, key=ordem), debug_map


def garantir_numerico(df, meses):
    if df is None:
        return None

    for m in meses:
        if m in df.columns:
            df[m] = pd.to_numeric(df[m], errors="coerce")
    return df


def garantir_colunas(df, cols, fill_value=0):
    """Garante que todas as colunas existam no dataframe."""
    if df is None:
        return None

    df = df.copy()
    for c in cols:
        if c not in df.columns:
            df[c] = fill_value
    return df


def colorir_valores(val):
    if isinstance(val, (int, float)):
        if val < 0:
            return "color:red;font-weight:bold;"
        if val > 0:
            return "color:green;font-weight:bold;"
    return ""


def formatar_tabela(df):
    out = df.copy()

    cols_num = out.select_dtypes(include="number").columns
    cols_txt = out.columns.difference(cols_num)

    if len(cols_num) > 0:
        out[cols_num] = out[cols_num].fillna(0)
    if len(cols_txt) > 0:
        out[cols_txt] = out[cols_txt].fillna("")

    styler = out.style.format(lambda x: f"{x:,.0f}".replace(",", "."), subset=cols_num)

    # Compatibilidade: pandas novos usam .map; antigos, .applymap
    if hasattr(styler, "map"):
        styler = styler.map(colorir_valores, subset=cols_num)
    else:
        styler = styler.applymap(colorir_valores, subset=cols_num)

    styler = (
        styler
        .set_properties(subset=cols_num, **{"text-align": "center"})
        .set_properties(subset=cols_txt, **{"text-align": "left"})
    )

    return styler



def colorir_percent(val):
    """
    Regras de cor para percentuais:
    <= 50%  -> vermelho
    51% a 79% -> amarelo
    >= 80% -> verde
    """
    if isinstance(val, (int, float)):
        if val <= 50:
            return "color:red;font-weight:bold;"
        elif val < 80:
            return "color:#d4a017;font-weight:bold;"   # amarelo mais legível no fundo claro/escuro
        else:
            return "color:green;font-weight:bold;"
    return ""



def formatar_tabela_percent(df):
    out = df.copy()

    cols_num = out.select_dtypes(include="number").columns
    cols_txt = out.columns.difference(cols_num)

    if len(cols_num) > 0:
        out[cols_num] = out[cols_num].fillna(0)
    if len(cols_txt) > 0:
        out[cols_txt] = out[cols_txt].fillna("")

    styler = out.style.format(lambda x: f"{x:.0f}%", subset=cols_num)

    if hasattr(styler, "map"):
        styler = styler.map(colorir_percent, subset=cols_num)
    else:
        styler = styler.applymap(colorir_percent, subset=cols_num)

    styler = (
        styler
        .set_properties(subset=cols_num, **{"text-align": "center"})
        .set_properties(subset=cols_txt, **{"text-align": "left"})
    )

    return styler


def normalizar_chaves(df, cols):
    if df is None:
        return None

    df = df.copy()
    for c in cols:
        if c in df.columns:
            df[c] = (
                df[c]
                .fillna("")
                .astype(str)
                .str.replace("\u00a0", " ", regex=False)
                .str.strip()
                .str.upper()
            )
    return df


def remover_linhas_sem_chave(df, cols_chave):
    """
    Remove linhas em que TODAS as colunas-chave estão vazias.
    Como você NÃO tem células mescladas, essa é a correção correta.
    """
    if df is None:
        return None

    df = df.copy()
    cols_existentes = [c for c in cols_chave if c in df.columns]

    if not cols_existentes:
        return df

    tmp = df[cols_existentes].copy()

    for c in cols_existentes:
        tmp[c] = (
            tmp[c]
            .fillna("")
            .astype(str)
            .str.replace("\u00a0", " ", regex=False)
            .str.strip()
        )

    mask_linha_vazia = tmp.eq("").all(axis=1)

    return df.loc[~mask_linha_vazia].copy()


def ocultar_linhas_sem_chave(df, cols_chave):
    """
    Defesa extra: evita exibir linhas totalmente vazias,
    caso alguma ainda escape.
    """
    if df is None:
        return df

    df = df.copy()
    cols_existentes = [c for c in cols_chave if c in df.columns]
    if not cols_existentes:
        return df

    tmp = df[cols_existentes].copy()
    for c in cols_existentes:
        tmp[c] = tmp[c].fillna("").astype(str).str.strip()

    mask = ~tmp.eq("").all(axis=1)
    return df.loc[mask].copy()


# =====================================================
# FUNÇÃO PRINCIPAL
# =====================================================
def gerar_passo1(xlsx_bytes, show_debug=False, visao="Request - Plan", incluir_outer=True):
    xls_original = pd.ExcelFile(io.BytesIO(xlsx_bytes), engine="openpyxl")

    plan = pd.read_excel(xls_original, "PLAN", engine="openpyxl")
    req = pd.read_excel(xls_original, "REQUEST", engine="openpyxl")

    fr = None
    if "F.RESPONSE" in xls_original.sheet_names:
        fr = pd.read_excel(xls_original, "F.RESPONSE", engine="openpyxl")

    if visao == "F.Response - Request" and fr is None:
        raise ValueError("Aba 'F.RESPONSE' não encontrada no Excel enviado.")

    # Detecta meses (união das abas disponíveis)
    meses_plan, map_plan = detectar_colunas_mes(plan)
    meses_req, map_req = detectar_colunas_mes(req)
    meses = list(dict.fromkeys(meses_plan + meses_req))

    map_fr = {}
    if fr is not None:
        meses_fr, map_fr = detectar_colunas_mes(fr)
        meses = list(dict.fromkeys(meses + meses_fr))

    if not meses:
        raise ValueError("Nenhuma coluna de mês encontrada.")

    # Garante que todos os dataframes tenham todos os meses
    plan = garantir_colunas(plan, meses, 0)
    req = garantir_colunas(req, meses, 0)
    fr = garantir_colunas(fr, meses, 0)

    # Força numérico
    plan = garantir_numerico(plan, meses)
    req = garantir_numerico(req, meses)
    fr = garantir_numerico(fr, meses)

    if show_debug:
        st.subheader("Diagnóstico PLAN")
        st.json(map_plan)
        st.subheader("Diagnóstico REQUEST")
        st.json(map_req)
        if fr is not None:
            st.subheader("Diagnóstico F.RESPONSE")
            st.json(map_fr)

    # =================================================
    # FILTROS — OPÇÃO C (DEPENDENTES / DINÂMICOS)
    # =================================================
    st.subheader("Filtros")

    frames = [plan, req]
    if fr is not None:
        frames.append(fr)
    union_df = pd.concat(frames, ignore_index=True, sort=False).copy()

    # garante string nas colunas de filtro
    for col in ["PRODUCT BRAND", "PRODUCT MARKET", "SITE", "PRODUCT NEED"]:
        if col in union_df.columns:
            union_df[col] = (
                union_df[col]
                .fillna("")
                .astype(str)
                .str.strip()
            )

    # inicializa estado
    for k in ["f_brand", "f_market", "f_site", "f_need"]:
        if k not in st.session_state:
            st.session_state[k] = []

    top1, top2 = st.columns([1.2, 8])
    with top1:
        if st.button("Limpar filtros", use_container_width=True):
            for k in ["f_brand", "f_market", "f_site", "f_need"]:
                st.session_state[k] = []
            st.rerun()

    with top2:
        st.caption("Os filtros são dependentes: cada seleção reduz as opções disponíveis nos demais.")

    def filtrar_base(df, brand=None, market=None, site=None, need=None):
        out = df.copy()

        if brand and "PRODUCT BRAND" in out.columns:
            out = out[out["PRODUCT BRAND"].isin(brand)]
        if market and "PRODUCT MARKET" in out.columns:
            out = out[out["PRODUCT MARKET"].isin(market)]
        if site and "SITE" in out.columns:
            out = out[out["SITE"].isin(site)]
        if need and "PRODUCT NEED" in out.columns:
            out = out[out["PRODUCT NEED"].isin(need)]

        return out

    def opcoes_validas(df, coluna):
        if coluna not in df.columns:
            return []
        vals = (
            df[coluna]
            .dropna()
            .astype(str)
            .str.strip()
        )
        vals = vals[vals != ""]
        return sorted(vals.unique().tolist())

    c1, c2, c3, c4 = st.columns([1.1, 1.6, 1.0, 1.5])

    # BRAND depende dos demais
    base_brand = filtrar_base(
        union_df,
        market=st.session_state["f_market"],
        site=st.session_state["f_site"],
        need=st.session_state["f_need"]
    )
    opts_brand = opcoes_validas(base_brand, "PRODUCT BRAND")
    st.session_state["f_brand"] = [v for v in st.session_state["f_brand"] if v in opts_brand]
    with c1:
        st.multiselect(
            "PRODUCT BRAND",
            options=opts_brand,
            default=st.session_state["f_brand"],
            key="f_brand",
            placeholder="Todos"
        )
        st.caption(f"{len(opts_brand)} opções")

    # MARKET depende dos demais
    base_market = filtrar_base(
        union_df,
        brand=st.session_state["f_brand"],
        site=st.session_state["f_site"],
        need=st.session_state["f_need"]
    )
    opts_market = opcoes_validas(base_market, "PRODUCT MARKET")
    st.session_state["f_market"] = [v for v in st.session_state["f_market"] if v in opts_market]
    with c2:
        st.multiselect(
            "PRODUCT MARKET",
            options=opts_market,
            default=st.session_state["f_market"],
            key="f_market",
            placeholder="Todos"
        )
        st.caption(f"{len(opts_market)} opções")

    # SITE depende dos demais
    base_site = filtrar_base(
        union_df,
        brand=st.session_state["f_brand"],
        market=st.session_state["f_market"],
        need=st.session_state["f_need"]
    )
    opts_site = opcoes_validas(base_site, "SITE")
    st.session_state["f_site"] = [v for v in st.session_state["f_site"] if v in opts_site]
    with c3:
        st.multiselect(
            "SITE",
            options=opts_site,
            default=st.session_state["f_site"],
            key="f_site",
            placeholder="Todos"
        )
        st.caption(f"{len(opts_site)} opções")

    # NEED depende dos demais
    base_need = filtrar_base(
        union_df,
        brand=st.session_state["f_brand"],
        market=st.session_state["f_market"],
        site=st.session_state["f_site"]
    )
    opts_need = opcoes_validas(base_need, "PRODUCT NEED")
    st.session_state["f_need"] = [v for v in st.session_state["f_need"] if v in opts_need]
    with c4:
        st.multiselect(
            "PRODUCT NEED",
            options=opts_need,
            default=st.session_state["f_need"],
            key="f_need",
            placeholder="Todos"
        )
        st.caption(f"{len(opts_need)} opções")

    def aplicar_filtros(df):
        if df is None:
            return None

        out = df.copy()

        if st.session_state["f_brand"] and "PRODUCT BRAND" in out.columns:
            out = out[out["PRODUCT BRAND"].fillna("").astype(str).isin(st.session_state["f_brand"])]

        if st.session_state["f_market"] and "PRODUCT MARKET" in out.columns:
            out = out[out["PRODUCT MARKET"].fillna("").astype(str).isin(st.session_state["f_market"])]

        if st.session_state["f_site"] and "SITE" in out.columns:
            out = out[out["SITE"].fillna("").astype(str).isin(st.session_state["f_site"])]

        if st.session_state["f_need"] and "PRODUCT NEED" in out.columns:
            out = out[out["PRODUCT NEED"].fillna("").astype(str).isin(st.session_state["f_need"])]

        return out

    plan = aplicar_filtros(plan)
    req = aplicar_filtros(req)
    fr = aplicar_filtros(fr)

    key_cols = ["SITE", "PRODUCT NEED", "PRODUCT SERIES", "PRODUCT BRAND", "PRODUCT MARKET"]

    # NORMALIZA AS CHAVES
    plan = normalizar_chaves(plan, key_cols)
    req = normalizar_chaves(req, key_cols)
    fr = normalizar_chaves(fr, key_cols)

    # REMOVE LINHAS TOTALMENTE SEM CHAVE (CORREÇÃO PRINCIPAL)
    plan = remover_linhas_sem_chave(plan, key_cols)
    req = remover_linhas_sem_chave(req, key_cols)
    fr = remover_linhas_sem_chave(fr, key_cols)

    if show_debug:
        st.subheader("Diagnóstico — linhas sem chave após limpeza")
        st.write("PLAN:", len(plan))
        st.write("REQUEST:", len(req))
        if fr is not None:
            st.write("F.RESPONSE:", len(fr))

    # =================================================
    # Seleção BASE e COMP conforme visão
    # =================================================
    if visao == "F.Response - Request":
        base_name, comp_name = "REQUEST", "F.RESPONSE"
        base_df, comp_df = req, fr
    else:
        base_name, comp_name = "PLAN", "REQUEST"
        base_df, comp_df = plan, req

    how_merge = "outer" if incluir_outer else "inner"

    # =================================================
    # TABELA DETALHADA — PRODUCT NEED + PRODUCT SERIES + BRAND + MARKET
    # =================================================
    grp_serie = [
        "SITE",
        "PRODUCT NEED",
        "PRODUCT SERIES",
        "PRODUCT BRAND",
        "PRODUCT MARKET",
    ]

    base_df = garantir_colunas(base_df, grp_serie + meses, 0)
    comp_df = garantir_colunas(comp_df, grp_serie + meses, 0)

    base_s = base_df[grp_serie + meses].groupby(grp_serie, dropna=False)[meses].sum().reset_index()
    comp_s = comp_df[grp_serie + meses].groupby(grp_serie, dropna=False)[meses].sum().reset_index()

    comp_s_merge = pd.merge(
        base_s, comp_s,
        on=grp_serie,
        how=how_merge,
        suffixes=(f"_{base_name}", f"_{comp_name}"),
        indicator=True
    )

    # preencher zero SOMENTE nas colunas numéricas de meses
    for m in meses:
        col_base = f"{m}_{base_name}"
        col_comp = f"{m}_{comp_name}"

        if col_base in comp_s_merge.columns:
            comp_s_merge[col_base] = comp_s_merge[col_base].fillna(0)

        if col_comp in comp_s_merge.columns:
            comp_s_merge[col_comp] = comp_s_merge[col_comp].fillna(0)

    for m in meses:
        comp_s_merge[m] = comp_s_merge[f"{m}_{comp_name}"] - comp_s_merge[f"{m}_{base_name}"]

    if show_debug:
        st.subheader("Diagnóstico do Merge - Comparativo Geral")
        st.write(comp_s_merge["_merge"].value_counts())

        st.write("Linhas só no BASE")
        st.dataframe(
            comp_s_merge[comp_s_merge["_merge"] == "left_only"][grp_serie + [f"{m}_{base_name}" for m in meses]],
            use_container_width=True
        )

        st.write("Linhas só no COMP")
        st.dataframe(
            comp_s_merge[comp_s_merge["_merge"] == "right_only"][grp_serie + [f"{m}_{comp_name}" for m in meses]],
            use_container_width=True
        )

    step1_serie = comp_s_merge[grp_serie + meses].copy()
    step1_serie = remover_linhas_sem_chave(step1_serie, grp_serie)
    step1_serie["TOTAL"] = step1_serie[meses].sum(axis=1)

    # Linha TOTAL GERAL (detalhado)
    total_s = {c: "TOTAL GERAL" for c in grp_serie}
    for m in meses:
        total_s[m] = step1_serie[m].sum()
    total_s["TOTAL"] = step1_serie["TOTAL"].sum()
    step1_serie = pd.concat([step1_serie, pd.DataFrame([total_s])], ignore_index=True)

    # =================================================
    # RESUMO POR SITE + BRAND + PRODUCT NEED (COMP - BASE)
    # =================================================
    grp_need = ["SITE", "PRODUCT BRAND", "PRODUCT NEED"]
    step1_serie_sem_total = step1_serie[step1_serie["SITE"].astype(str).str.upper() != "TOTAL GERAL"].copy()

    step1_need = (
        step1_serie_sem_total[grp_need + meses]
        .groupby(grp_need, dropna=False)[meses]
        .sum()
        .reset_index()
    )

    step1_need = remover_linhas_sem_chave(step1_need, grp_need)
    step1_need["TOTAL"] = step1_need[meses].sum(axis=1)

    total_n = {c: "TOTAL GERAL" for c in grp_need}
    for m in meses:
        total_n[m] = step1_need[m].sum()
    total_n["TOTAL"] = step1_need["TOTAL"].sum()

    step1_need = pd.concat([step1_need, pd.DataFrame([total_n])], ignore_index=True)

    # =================================================
    # RESUMO FINAL POR SITE + BRAND + PRODUCT NEED (SOMENTE COMP)
    # =================================================
    comp_only_need = (
        comp_df[grp_need + meses]
        .groupby(grp_need, dropna=False)[meses]
        .sum()
        .reset_index()
    )

    comp_only_need = remover_linhas_sem_chave(comp_only_need, grp_need)
    comp_only_need["TOTAL"] = comp_only_need[meses].sum(axis=1)

    total_comp = {c: "TOTAL GERAL" for c in grp_need}
    for m in meses:
        total_comp[m] = comp_only_need[m].sum()
    total_comp["TOTAL"] = comp_only_need["TOTAL"].sum()

    comp_only_need = pd.concat([comp_only_need, pd.DataFrame([total_comp])], ignore_index=True)

    # =================================================
    # NOVA TABELA — % de Atendimento (por Quarter)
    # Definição:
    # - Request - Plan: PLAN / REQUEST
    # - F.Response - Request: F.RESPONSE / REQUEST
    # =================================================
    def _mes_to_quarter(m_alias: str) -> str:
        mm = _normalize_header(m_alias).split('/')[0]
        if mm in ['jan', 'fev', 'mar']:
            return 'Q1'
        if mm in ['abr', 'mai', 'jun']:
            return 'Q2'
        if mm in ['jul', 'ago', 'set']:
            return 'Q3'
        return 'Q4'

    quarter_months = {'Q1': [], 'Q2': [], 'Q3': [], 'Q4': []}
    for m in meses:
        q = _mes_to_quarter(str(m))
        quarter_months[q].append(m)

    # demanda sempre é REQUEST
    demand_df = req
    # oferta depende da visão
    supply_df = fr if visao == 'F.Response - Request' else plan

    grp_att = ["SITE", "PRODUCT NEED", "PRODUCT SERIES"]

    demand_df = garantir_colunas(demand_df, grp_att + meses, 0)
    supply_df = garantir_colunas(supply_df, grp_att + meses, 0)

    dem_g = demand_df[grp_att + meses].groupby(grp_att, dropna=False)[meses].sum().reset_index()
    sup_g = supply_df[grp_att + meses].groupby(grp_att, dropna=False)[meses].sum().reset_index()

    dem_g = remover_linhas_sem_chave(dem_g, grp_att)
    sup_g = remover_linhas_sem_chave(sup_g, grp_att)

    how_att = 'outer'
    att = pd.merge(dem_g, sup_g, on=grp_att, how=how_att, suffixes=("_DEM", "_SUP"))

    for m in meses:
        dem_col = f"{m}_DEM"
        sup_col = f"{m}_SUP"
        if dem_col in att.columns:
            att[dem_col] = att[dem_col].fillna(0)
        if sup_col in att.columns:
            att[sup_col] = att[sup_col].fillna(0)

    for q, mlist in quarter_months.items():
        if not mlist:
            att[q] = 100.0
            continue

        dem_q = sum(att[f"{m}_DEM"] for m in mlist)
        sup_q = sum(att[f"{m}_SUP"] for m in mlist)

        att[q] = (sup_q / dem_q * 100).where(dem_q > 0, 100.0)

    dem_tot = sum(att[f"{m}_DEM"] for m in meses)
    sup_tot = sum(att[f"{m}_SUP"] for m in meses)
    att["TOTAL"] = (sup_tot / dem_tot * 100).where(dem_tot > 0, 100.0)

    for c in ['Q1', 'Q2', 'Q3', 'Q4', 'TOTAL']:
        att[c] = att[c].clip(lower=0, upper=100)

    df_atendimento = att[grp_att + ['Q1', 'Q2', 'Q3', 'Q4', 'TOTAL']].copy()
    df_atendimento = remover_linhas_sem_chave(df_atendimento, grp_att)

    total_att = {c: 'TOTAL GERAL' for c in grp_att}
    for q, mlist in quarter_months.items():
        if not mlist:
            total_att[q] = 100.0
            continue

        dem_q = sum(att[f"{m}_DEM"] for m in mlist).sum()
        sup_q = sum(att[f"{m}_SUP"] for m in mlist).sum()

        total_att[q] = (sup_q / dem_q * 100) if dem_q > 0 else 100.0
        total_att[q] = max(0.0, min(100.0, total_att[q]))

    dem_t = dem_tot.sum()
    sup_t = sup_tot.sum()
    total_att['TOTAL'] = (sup_t / dem_t * 100) if dem_t > 0 else 100.0
    total_att['TOTAL'] = max(0.0, min(100.0, total_att['TOTAL']))

    df_atendimento = pd.concat([df_atendimento, pd.DataFrame([total_att])], ignore_index=True)

    # =================================================
    # ADIÇÃO — PRODUCT · FC · DELTA MENSAL (COMP - BASE)
    # Mantém exatamente o padrão de colunas das abas originais
    # =================================================
    if "DEMAND TYPE" in base_df.columns:
        base_fc = base_df[base_df["DEMAND TYPE"] == "FC"].copy()
    else:
        base_fc = base_df.iloc[0:0].copy()

    if "DEMAND TYPE" in comp_df.columns:
        comp_fc = comp_df[comp_df["DEMAND TYPE"] == "FC"].copy()
    else:
        comp_fc = comp_df.iloc[0:0].copy()

    month_positions = [plan.columns.get_loc(c) for c in meses if c in plan.columns]
    if month_positions:
        first_month_pos = min(month_positions)
        meta_cols = list(plan.columns[:first_month_pos])
    else:
        month_positions_b = [base_df.columns.get_loc(c) for c in meses if c in base_df.columns]
        first_month_pos = min(month_positions_b) if month_positions_b else len(base_df.columns)
        meta_cols = list(base_df.columns[:first_month_pos])

    base_fc = garantir_colunas(base_fc, meta_cols + meses, 0)
    comp_fc = garantir_colunas(comp_fc, meta_cols + meses, 0)

    # remove linhas totalmente vazias nas colunas-meta
    base_fc = remover_linhas_sem_chave(base_fc, [c for c in meta_cols if c in base_fc.columns] or meta_cols)
    comp_fc = remover_linhas_sem_chave(comp_fc, [c for c in meta_cols if c in comp_fc.columns] or meta_cols)

    base_p = base_fc[meta_cols + meses].groupby(meta_cols, dropna=False)[meses].sum().reset_index()
    comp_p = comp_fc[meta_cols + meses].groupby(meta_cols, dropna=False)[meses].sum().reset_index()

    comp_p_merge = pd.merge(
        base_p, comp_p,
        on=meta_cols,
        how=how_merge,
        suffixes=(f"_{base_name}", f"_{comp_name}")
    )

    for m in meses:
        col_base = f"{m}_{base_name}"
        col_comp = f"{m}_{comp_name}"
        if col_base in comp_p_merge.columns:
            comp_p_merge[col_base] = comp_p_merge[col_base].fillna(0)
        if col_comp in comp_p_merge.columns:
            comp_p_merge[col_comp] = comp_p_merge[col_comp].fillna(0)

    step1_product_fc = comp_p_merge[meta_cols].copy()
    for m in meses:
        step1_product_fc[m] = comp_p_merge[f"{m}_{comp_name}"] - comp_p_merge[f"{m}_{base_name}"]

    step1_product_fc = remover_linhas_sem_chave(step1_product_fc, meta_cols)
    step1_product_fc["TOTAL"] = step1_product_fc[meses].sum(axis=1)

    total_prod = {c: "TOTAL GERAL" for c in meta_cols}
    for m in meses:
        total_prod[m] = step1_product_fc[m].sum()
    total_prod["TOTAL"] = step1_product_fc["TOTAL"].sum()
    step1_product_fc = pd.concat([step1_product_fc, pd.DataFrame([total_prod])], ignore_index=True)

    
# =================================================
    # EXPORTAR EXCEL (FORMATADO)
    # =================================================
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    buf_out = io.BytesIO()
    with pd.ExcelWriter(buf_out, engine="openpyxl") as writer:
        for sheet in xls_original.sheet_names:
            pd.read_excel(xls_original, sheet, engine="openpyxl").to_excel(
                writer,
                sheet_name=sheet,
                index=False
            )

        abas = {
            "Step1_Comparativo_Serie": step1_serie,
            "Step1_Comparativo_Need": step1_need,
            f"Resumo_{comp_name}_Product_Need": comp_only_need,
            "Atendimento_%_Quarter": df_atendimento,
            f"{comp_name} x {base_name} FC - Produto Mensal": step1_product_fc,
        }

        for nome, df in abas.items():
            sheet_name = nome[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.book[sheet_name]

            # ============================================
            # FORMATAÇÃO ESPECIAL PARA ATENDIMENTO (%)
            # ============================================
            if sheet_name == "Atendimento_%_Quarter":
                perc_cols = ["Q1", "Q2", "Q3", "Q4", "TOTAL"]

                for pc in perc_cols:
                    if pc in df.columns:
                        col_idx = df.columns.get_loc(pc) + 1

                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                            cell = row[col_idx - 1]

                            if isinstance(cell.value, (int, float)):
                                valor_original = cell.value  # ex.: 45, 72, 88
                                cell.value = valor_original / 100.0
                                cell.number_format = "0%"

                                if valor_original <= 50:
                                    cell.font = Font(color="FF0000", bold=True)   # vermelho
                                elif valor_original < 80:
                                    cell.font = Font(color="FFC000", bold=True)   # amarelo
                                else:
                                    cell.font = Font(color="008000", bold=True)   # verde

            # ============================================
            # FORMATAÇÃO DAS DEMAIS ABAS NUMÉRICAS
            # ============================================
            cols_num = df.select_dtypes(include="number").columns
            idx_cols = [df.columns.get_loc(c) + 1 for c in cols_num]

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for idx in idx_cols:
                    cell = row[idx - 1]

                    if isinstance(cell.value, (int, float)):
                        if sheet_name != "Atendimento_%_Quarter":
                            cell.number_format = "#,##0"

                            if cell.value < 0:
                                cell.font = Font(color="FF0000", bold=True)
                            elif cell.value > 0:
                                cell.font = Font(color="008000", bold=True)

                if str(row[0].value).upper() == "TOTAL GERAL":
                    for cell in row:
                        cell.font = Font(bold=True)

            # Ajusta largura das colunas
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    if cell.value is None:
                        continue
                    max_len = max(max_len, len(str(cell.value)))

                ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    return buf_out.getvalue(), step1_serie, step1_need, comp_only_need, df_atendimento


# =====================================================
# UI
# =====================================================
uploaded = st.file_uploader("Envie o Excel (PLAN, REQUEST e opcionalmente F.RESPONSE)", type=["xlsx"])

col1, col2, col3 = st.columns([2, 2, 3])
with col1:
    visao = st.radio("Visão", ["Request - Plan", "F.Response - Request"], horizontal=True)

with col3:
    debug = st.checkbox("Exibir diagnóstico", value=False)


if uploaded:
    try:
        excel_out, df_serie, df_need, df_comp_need, df_atend = gerar_passo1(
            uploaded.read(),
            show_debug=debug,
            visao=visao,
        )

        # Defesa extra na exibição
        df_serie_view = ocultar_linhas_sem_chave(
            df_serie,
            ["SITE", "PRODUCT NEED", "PRODUCT SERIES", "PRODUCT BRAND", "PRODUCT MARKET"]
        )

        df_need_view = ocultar_linhas_sem_chave(
            df_need,
            ["SITE", "PRODUCT BRAND", "PRODUCT NEED"]
        )

        df_comp_need_view = ocultar_linhas_sem_chave(
            df_comp_need,
            ["SITE", "PRODUCT BRAND", "PRODUCT NEED"]
        )

        df_atend_view = ocultar_linhas_sem_chave(
            df_atend,
            ["SITE", "PRODUCT NEED", "PRODUCT SERIES"]
        )

        st.subheader("Comparativo por SITE + BRAND + PRODUCT NEED")
        st.dataframe(formatar_tabela(df_need_view), use_container_width=True)

        st.subheader("Comparativo Geral")
        st.dataframe(formatar_tabela(df_serie_view), use_container_width=True)        

        st.subheader("Resumo final por SITE + BRAND + PRODUCT NEED")
        st.dataframe(formatar_tabela(df_comp_need_view), use_container_width=True)

        # MOSTRAR APENAS NA VISÃO F.Response - Request
        if visao == "F.Response - Request":
            st.subheader("% de Atendimento (por Quarter)")
            st.dataframe(formatar_tabela_percent(df_atend_view), use_container_width=True)

        nome_saida = f"saida_step1_{visao.replace(' ', '_').replace('.', '')}_.xlsx"
        st.download_button(
            "⬇️ Baixar Excel",
            data=excel_out,
            file_name=nome_saida
        )

    except Exception as e:
        st.error("Erro ao processar o arquivo")
        st.exception(e)
else:
    st.info("Faça upload do Excel para iniciar.")

st.markdown(
    """
    <style>
    :root {
        --sidebar-width: 21rem;
    }

    .footer-bar {
        position: fixed;
        bottom: 0;
        left: var(--sidebar-width);
        width: calc(100% - var(--sidebar-width));
        background-color: rgba(14, 17, 23, 0.95);
        color: #ccc;
        font-size: 0.75rem;
        text-align: center;
        padding: 8px 0;
        z-index: 999;
        border-top: 1px solid #333;
        backdrop-filter: blur(4px);
    }
    </style>

    <div class="footer-bar">
        Aplicação desenvolvida para suporte às análises do time MPS • Versão 1.0 — Jeferson Santana / Copilot
    </div>
    """,
    unsafe_allow_html=True
)
